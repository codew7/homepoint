import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter

# ── Datos INDEC (IPC variación anual %) ──────────────────────────────────────
datos = [
    (2016, 40.3,  "Retorno de estadísticas oficiales tras intervención 2007-2015"),
    (2017, 24.8,  "Leve desaceleración; metas de inflación no cumplidas"),
    (2018, 47.6,  "Crisis cambiaria; devaluación del peso y rescate FMI"),
    (2019, 53.8,  "Controles de capital (cepo); incertidumbre electoral"),
    (2020, 36.1,  "Pandemia COVID-19; controles de precios parciales"),
    (2021, 50.9,  "Emisión monetaria post-pandemia; desequilibrio fiscal"),
    (2022, 94.8,  "Sequía, guerra Ucrania; crisis de deuda y cambiaria"),
    (2023, 211.4, "Devaluación diciembre 2023; cambio de gobierno"),
    (2024, 117.8, "Ajuste fiscal Milei; desaceleración gradual mensual"),
    (2025, 47.3,  "Continuidad plan de estabilización; baja sostenida"),
]

AÑOS  = [d[0] for d in datos]
TASAS = [d[1] for d in datos]

# ── Paleta ────────────────────────────────────────────────────────────────────
AZUL_OSCURO = "1F3864"
AZUL_MEDIO  = "2E75B6"
AZUL_CLARO  = "BDD7EE"
GRIS_CLARO  = "F2F2F2"
BLANCO      = "FFFFFF"
NEGRO       = "000000"
ROJO        = "C00000"
NARANJA     = "ED7D31"
VERDE       = "70AD47"

def color_tasa(t):
    return ROJO if t >= 100 else (NARANJA if t >= 50 else VERDE)

def borde_fino():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inflación Argentina"

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 50

# ── Encabezado ────────────────────────────────────────────────────────────────
ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "INFLACIÓN ARGENTINA — ÚLTIMOS 10 AÑOS"
c.font = Font(name="Calibri", bold=True, size=18, color=BLANCO)
c.fill = PatternFill("solid", fgColor=AZUL_OSCURO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:D2")
c = ws["A2"]
c.value = "Variación anual del IPC (Índice de Precios al Consumidor) — Fuente: INDEC"
c.font = Font(name="Calibri", italic=True, size=10, color=AZUL_OSCURO)
c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 8

# ── Cabecera tabla ────────────────────────────────────────────────────────────
headers = ["Año", "Inflación Anual (%)", "Nivel de Impacto", "Contexto Económico"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL_MEDIO)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borde_fino()
ws.row_dimensions[4].height = 28

# ── Filas de datos ────────────────────────────────────────────────────────────
for i, (año, tasa, nota) in enumerate(datos):
    fila = 5 + i
    bg = GRIS_CLARO if i % 2 == 0 else BLANCO
    color = color_tasa(tasa)

    c = ws.cell(row=fila, column=1, value=año)
    c.font = Font(name="Calibri", bold=True, size=11, color=AZUL_OSCURO)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borde_fino()

    c = ws.cell(row=fila, column=2, value=tasa / 100)
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", bold=True, size=12, color=color)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borde_fino()

    if tasa >= 100:
        nivel, nc = "CRITICO", ROJO
    elif tasa >= 50:
        nivel, nc = "ALTO", NARANJA
    else:
        nivel, nc = "MODERADO", VERDE
    c = ws.cell(row=fila, column=3, value=nivel)
    c.font = Font(name="Calibri", bold=True, size=10, color=nc)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borde_fino()

    c = ws.cell(row=fila, column=4, value=nota)
    c.font = Font(name="Calibri", size=10, color=NEGRO)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = borde_fino()
    ws.row_dimensions[fila].height = 22

# ── Fila promedio ─────────────────────────────────────────────────────────────
fila_prom = 5 + len(datos)
prom = sum(TASAS) / len(TASAS)

for col, val in enumerate([None, prom / 100, None, None], 1):
    c = ws.cell(row=fila_prom, column=col)
    c.fill = PatternFill("solid", fgColor=AZUL_OSCURO)
    c.border = borde_fino()
    c.alignment = Alignment(horizontal="center", vertical="center")
    if col == 1:
        c.value = "Promedio"
        c.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    elif col == 2:
        c.value = prom / 100
        c.number_format = "0.0%"
        c.font = Font(name="Calibri", bold=True, size=12, color=BLANCO)

ws.merge_cells(f"C{fila_prom}:D{fila_prom}")
c = ws.cell(row=fila_prom, column=3)
c.value = f"Promedio de inflación anual 2016–2025 = {prom:.1f}%"
c.font = Font(name="Calibri", italic=True, size=10, color=BLANCO)
c.fill = PatternFill("solid", fgColor=AZUL_OSCURO)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = borde_fino()
ws.row_dimensions[fila_prom].height = 22

# ── Columnas auxiliares para el gráfico (F, G, H — ocultas después) ───────────
for i, (año, tasa, _) in enumerate(datos):
    ws.cell(row=5 + i, column=6, value=año)
    ws.cell(row=5 + i, column=7, value=tasa)
    ws.cell(row=5 + i, column=8, value=prom)   # línea promedio plana

# ── Gráfico de barras ─────────────────────────────────────────────────────────
bar = BarChart()
bar.type = "col"
bar.title = "Inflación Argentina 2016–2025 (% anual)"
bar.y_axis.title = "Variación anual (%)"
bar.x_axis.title = "Año"
bar.style = 10
bar.width = 24
bar.height = 14
bar.grouping = "clustered"

data_ref = Reference(ws, min_col=7, min_row=5, max_row=14)
cats_ref = Reference(ws, min_col=6, min_row=5, max_row=14)
bar.add_data(data_ref)
bar.set_categories(cats_ref)
bar.series[0].title = SeriesLabel(v="IPC anual (%)")

# Color individual por barra
COLOR_MAP = [color_tasa(t) for t in TASAS]
for idx, hex_color in enumerate(COLOR_MAP):
    pt = DataPoint(idx=idx)
    pt.spPr = GraphicalProperties(solidFill=hex_color)
    bar.series[0].dPt.append(pt)

bar.series[0].spPr = GraphicalProperties(solidFill=AZUL_MEDIO)

# Etiquetas de valor sobre las barras
bar.series[0].dLbls = DataLabelList()
bar.series[0].dLbls.showVal = True
bar.series[0].dLbls.showLegendKey = False
bar.series[0].dLbls.showCatName = False
bar.series[0].dLbls.showSerName = False

# ── Línea de promedio ─────────────────────────────────────────────────────────
line = LineChart()
line_ref = Reference(ws, min_col=8, min_row=5, max_row=14)
line.add_data(line_ref)
line.series[0].title = SeriesLabel(v=f"Promedio ({prom:.1f}%)")
line.series[0].graphicalProperties.line.solidFill = AZUL_OSCURO
line.series[0].graphicalProperties.line.width = 25000
line.series[0].smooth = True

bar += line
ws.add_chart(bar, "A17")

# ── Leyenda lateral ───────────────────────────────────────────────────────────
ws.column_dimensions["F"].width = 22
leyenda = [
    ("LEYENDA",        AZUL_OSCURO, True),
    ("MODERADO < 50%", VERDE,       False),
    ("ALTO  50–99%",   NARANJA,     False),
    ("CRITICO >= 100%",ROJO,        False),
]
for j, (txt, col, bold) in enumerate(leyenda):
    c = ws.cell(row=17 + j, column=6, value=txt)
    c.font = Font(name="Calibri", size=9, bold=bold, color=BLANCO if bold else col)
    if bold:
        c.fill = PatternFill("solid", fgColor=AZUL_OSCURO)
    c.alignment = Alignment(horizontal="left", vertical="center")

# ── Pie de página ─────────────────────────────────────────────────────────────
fila_pie = 33
ws.merge_cells(f"A{fila_pie}:D{fila_pie}")
c = ws.cell(row=fila_pie, column=1)
c.value = (
    "Fuente: INDEC — Instituto Nacional de Estadística y Censos de Argentina  |  "
    "Datos: IPC Nacional  |  2025 estimado sujeto a revisión  |  Elaboración propia 2026"
)
c.font = Font(name="Calibri", italic=True, size=8, color="808080")
c.alignment = Alignment(horizontal="center")

# ── Ocultar columnas auxiliares ───────────────────────────────────────────────
for col in ("G", "H"):
    ws.column_dimensions[col].hidden = True

# ── Configuración de página ───────────────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

# ── Guardar ───────────────────────────────────────────────────────────────────
output = "/home/user/homepoint/inflacion_argentina_2016_2025.xlsx"
wb.save(output)
print(f"Archivo generado correctamente: {output}")
